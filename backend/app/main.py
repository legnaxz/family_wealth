from datetime import date, datetime, timedelta, timezone
from collections import defaultdict
import hashlib
import secrets
import os
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from sqlalchemy import select, func
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .database import get_db, Base, engine
from .models import (
    User,
    Household,
    HouseholdMember,
    Account,
    Asset,
    Liability,
    Valuation,
    Transaction,
    InvitationToken,
    NetWorthSnapshot,
    AuditLog,
)
from .schemas import UserCreate, LoginIn, TokenOut, HouseholdCreate, AccountCreate, AssetCreate, LiabilityCreate, ValuationCreate
from .security import (
    hash_password,
    verify_password,
    create_access_token,
    decode_access_token,
    generate_totp_secret,
    verify_totp,
    totp_uri,
)

app = FastAPI(title="Family Wealth MVP")
security = HTTPBearer(auto_error=False)
AUTH_DISABLED = os.getenv("AUTH_DISABLED", "false").lower() in {"1", "true", "yes"}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ROLE_RANK = {"viewer": 1, "member": 2, "admin": 3, "owner": 4}


def ensure_local_household(db: Session) -> tuple[User, Household]:
    # local convenience: auto-create schema if migrations weren't run yet
    # (sqlite race-safe: ignore "already exists" during concurrent bootstrap)
    try:
        Base.metadata.create_all(bind=engine)
    except Exception:
        pass

    demo = db.scalar(select(User).where(User.email == "demo@local"))
    if not demo:
        demo = User(email="demo@local", password_hash=hash_password("demo"))
        db.add(demo)
        db.commit()
        db.refresh(demo)

    household = db.get(Household, 1)
    if not household:
        household = Household(id=1, name="우리집", owner_user_id=demo.id)
        db.add(household)
        db.commit()
        db.refresh(household)

    member = db.scalar(select(HouseholdMember).where(HouseholdMember.household_id == household.id, HouseholdMember.user_id == demo.id))
    if not member:
        db.add(HouseholdMember(household_id=household.id, user_id=demo.id, role="owner"))
        db.commit()

    return demo, household


def excel_serial_to_date(value) -> date:
    if isinstance(value, date):
        return value
    try:
        n = float(value)
        return (datetime(1899, 12, 30) + timedelta(days=n)).date()
    except Exception:
        return date.fromisoformat(str(value))


def norm_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return excel_serial_to_date(v)


def get_current_user(
    cred: HTTPAuthorizationCredentials | None = Depends(security),
    db: Session = Depends(get_db),
) -> User:
    if AUTH_DISABLED:
        demo, _ = ensure_local_household(db)
        return demo

    if not cred:
        raise HTTPException(401, "missing access token")
    try:
        user_id = decode_access_token(cred.credentials)
    except ValueError:
        raise HTTPException(401, "invalid access token")
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(401, "user not found")
    return user


def get_household_role(db: Session, household_id: int, user_id: int) -> str | None:
    m = db.scalar(
        select(HouseholdMember).where(
            HouseholdMember.household_id == household_id,
            HouseholdMember.user_id == user_id,
        )
    )
    return m.role if m else None


def require_household_role(db: Session, household_id: int, user_id: int, minimum_role: str) -> str:
    if AUTH_DISABLED:
        return "owner"
    role = get_household_role(db, household_id, user_id)
    if not role:
        raise HTTPException(403, "not a household member")
    if ROLE_RANK.get(role, 0) < ROLE_RANK.get(minimum_role, 999):
        raise HTTPException(403, f"requires role >= {minimum_role}")
    return role


def audit(db: Session, household_id: int, actor_user_id: int, action: str, target_type: str, target_id: str, detail: str | None = None):
    db.add(
        AuditLog(
            household_id=household_id,
            actor_user_id=actor_user_id,
            action=action,
            target_type=target_type,
            target_id=target_id,
            detail=detail,
        )
    )


@app.get("/health")
def health(db: Session = Depends(get_db)):
    if AUTH_DISABLED:
        _, h = ensure_local_household(db)
        tx_count = db.scalar(select(func.count()).select_from(Transaction).where(Transaction.household_id == h.id)) or 0
        snap_count = db.scalar(select(func.count()).select_from(NetWorthSnapshot).where(NetWorthSnapshot.household_id == h.id)) or 0
        return {"ok": True, "authDisabled": True, "householdId": h.id, "transactions": int(tx_count), "snapshots": int(snap_count)}
    return {"ok": True}


@app.post("/local/bootstrap")
def local_bootstrap(db: Session = Depends(get_db)):
    demo, household = ensure_local_household(db)
    return {"user": demo.email, "household_id": household.id, "household_name": household.name}


@app.post("/auth/register", response_model=TokenOut)
def register(payload: UserCreate, db: Session = Depends(get_db)):
    exists = db.scalar(select(User).where(User.email == payload.email))
    if exists:
        raise HTTPException(400, "email already exists")
    user = User(email=payload.email, password_hash=hash_password(payload.password))
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"access_token": create_access_token(user.id)}


@app.post("/auth/login", response_model=TokenOut)
def login(payload: LoginIn, db: Session = Depends(get_db)):
    user = db.scalar(select(User).where(User.email == payload.email))
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(401, "invalid credentials")
    if user.otp_enabled:
        if not payload.otp_code or not user.otp_secret or not verify_totp(user.otp_secret, payload.otp_code):
            raise HTTPException(401, "2fa code required or invalid")
    return {"access_token": create_access_token(user.id)}


@app.post("/auth/2fa/setup")
def setup_2fa(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if user.otp_enabled and user.otp_secret:
        return {"secret": user.otp_secret, "otpauth_uri": totp_uri(user.otp_secret, user.email)}
    secret = generate_totp_secret()
    user.otp_secret = secret
    db.commit()
    return {"secret": secret, "otpauth_uri": totp_uri(secret, user.email)}


@app.post("/auth/2fa/enable")
def enable_2fa(code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not user.otp_secret:
        raise HTTPException(400, "2fa setup required")
    if not verify_totp(user.otp_secret, code):
        raise HTTPException(400, "invalid otp code")
    user.otp_enabled = True
    db.commit()
    return {"ok": True, "otp_enabled": True}


@app.post("/households")
def create_household(payload: HouseholdCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    h = Household(name=payload.name, owner_user_id=user.id)
    db.add(h)
    db.commit()
    db.refresh(h)

    db.add(HouseholdMember(household_id=h.id, user_id=user.id, role="owner"))
    audit(db, h.id, user.id, "create", "household", str(h.id))
    db.commit()
    return {"id": h.id, "name": h.name}


@app.post("/households/{household_id}/members")
def add_household_member(household_id: int, email: str, role: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if role not in ROLE_RANK:
        raise HTTPException(400, "invalid role")
    require_household_role(db, household_id, user.id, "admin")
    target = db.scalar(select(User).where(User.email == email))
    if not target:
        raise HTTPException(404, "target user not found")

    existing = db.scalar(select(HouseholdMember).where(HouseholdMember.household_id == household_id, HouseholdMember.user_id == target.id))
    if existing:
        existing.role = role
    else:
        db.add(HouseholdMember(household_id=household_id, user_id=target.id, role=role))
    audit(db, household_id, user.id, "upsert", "household_member", f"{target.id}:{role}")
    db.commit()
    return {"ok": True}


@app.post("/households/{household_id}/invite-tokens")
def create_invite_token(
    household_id: int,
    role: str = "member",
    expires_hours: int = 72,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_household_role(db, household_id, user.id, "admin")
    if role not in ROLE_RANK:
        raise HTTPException(400, "invalid role")
    if role == "owner":
        raise HTTPException(400, "owner invite not allowed")
    token = secrets.token_urlsafe(24)
    row = InvitationToken(
        token=token,
        household_id=household_id,
        role=role,
        created_by_user_id=user.id,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=expires_hours),
    )
    db.add(row)
    audit(db, household_id, user.id, "create", "invitation_token", token, detail=f"role={role}")
    db.commit()
    return {"token": token, "role": role, "expiresHours": expires_hours}


@app.post("/households/join")
def join_household(token: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.get(InvitationToken, token)
    if not inv:
        raise HTTPException(404, "invalid invite token")
    if inv.used:
        raise HTTPException(400, "invite already used")
    if inv.expires_at.replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        raise HTTPException(400, "invite expired")

    existing = db.scalar(select(HouseholdMember).where(HouseholdMember.household_id == inv.household_id, HouseholdMember.user_id == user.id))
    if existing:
        existing.role = inv.role
    else:
        db.add(HouseholdMember(household_id=inv.household_id, user_id=user.id, role=inv.role))

    inv.used = True
    inv.used_by_user_id = user.id
    audit(db, inv.household_id, user.id, "join", "invitation_token", token, detail=f"role={inv.role}")
    db.commit()
    return {"ok": True, "household_id": inv.household_id, "role": inv.role}


@app.post("/accounts")
def create_account(payload: AccountCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, payload.household_id, user.id, "member")
    row = Account(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    audit(db, row.household_id, user.id, "create", "account", str(row.id))
    db.commit()
    return {"id": row.id}


@app.post("/assets")
def create_asset(payload: AssetCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, payload.household_id, user.id, "member")
    row = Asset(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    audit(db, row.household_id, user.id, "create", "asset", str(row.id))
    db.commit()
    return {"id": row.id}


@app.post("/liabilities")
def create_liability(payload: LiabilityCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, payload.household_id, user.id, "member")
    row = Liability(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    audit(db, row.household_id, user.id, "create", "liability", str(row.id))
    db.commit()
    return {"id": row.id}


@app.post("/valuations")
def create_valuation(payload: ValuationCreate, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, payload.household_id, user.id, "member")
    row = Valuation(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    audit(db, row.household_id, user.id, "create", "valuation", str(row.id))
    db.commit()
    return {"id": row.id}


@app.post("/imports/xlsx")
def import_xlsx(household_id: int = 1, file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "member")
    wb = load_workbook(file.file, data_only=True)

    imported = 0
    skipped_duplicates = 0
    imported_assets = 0
    imported_liabilities = 0

    # sheet1(뱅샐현황) 재무현황 파싱: 자산/부채 현재가치 적재
    def _num(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        if s in {"", "-"}:
            return 0.0
        try:
            return float(s)
        except Exception:
            return 0.0

    if "뱅샐현황" in wb.sheetnames:
        ws0 = wb["뱅샐현황"]
        as_of = date.today()

        in_financial_section = False
        for row in ws0.iter_rows(min_row=1, values_only=True):
            vals = list(row)
            line = " ".join(str(v) for v in vals if v is not None)

            if "3.재무현황" in line:
                in_financial_section = True
                continue
            if "4.보험현황" in line and in_financial_section:
                in_financial_section = False
                break
            if not in_financial_section:
                continue

            # 실제 열 구조: B(자산항목) C(자산명) E(자산금액), F(부채항목) G(부채명) I(부채금액)
            if len(vals) < 9:
                continue

            asset_type = str(vals[1]).strip() if vals[1] is not None else ""
            asset_name = str(vals[2]).strip() if vals[2] is not None else ""
            liab_type = str(vals[5]).strip() if vals[5] is not None else ""
            liab_name = str(vals[6]).strip() if vals[6] is not None else ""
            asset_amount = _num(vals[4])
            liab_amount = _num(vals[8])

            # 헤더/합계/안내행 제외
            if asset_name in {"", "상품명", "총자산"}:
                asset_amount = 0
            if liab_name in {"", "상품명", "총부채"}:
                liab_amount = 0
            if "데이터를 내보낸" in asset_type:
                continue

            if asset_name and asset_amount != 0:
                a = db.scalar(select(Asset).where(Asset.household_id == household_id, Asset.name == asset_name))
                if not a:
                    a = Asset(household_id=household_id, name=asset_name, category="sheet1")
                    db.add(a)
                    db.flush()
                db.add(Valuation(household_id=household_id, asset_id=a.id, as_of_date=as_of, amount=asset_amount))
                imported_assets += 1

            if liab_name and liab_amount != 0:
                l = db.scalar(select(Liability).where(Liability.household_id == household_id, Liability.name == liab_name))
                if not l:
                    l = Liability(household_id=household_id, name=liab_name)
                    db.add(l)
                    db.flush()
                db.add(Valuation(household_id=household_id, liability_id=l.id, as_of_date=as_of, amount=liab_amount))
                imported_liabilities += 1

        # sheet1 결과는 먼저 커밋(이후 거래중복 등으로 롤백되지 않게)
        db.commit()

    if "가계부 내역" in wb.sheetnames:
        ws = wb["가계부 내역"]
        seen_hashes: set[str] = set()
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 날짜,시간,타입,대분류,소분류,내용,금액,화폐,결제수단,메모
            if not row or row[0] is None or row[2] is None or row[6] is None:
                continue
            tx_date = excel_serial_to_date(row[0])
            tx_type = str(row[2]).strip()
            category = str(row[3]).strip() if row[3] else None
            subcategory = str(row[4]).strip() if row[4] else None
            content = str(row[5]).strip() if row[5] else None
            amount = float(row[6])
            currency = str(row[7]).strip() if row[7] else "KRW"
            payment_method = str(row[8]).strip() if row[8] else None

            raw_key = f"{tx_date}|{tx_type}|{category}|{subcategory}|{content}|{amount}|{currency}|{payment_method}"
            tx_hash = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
            if tx_hash in seen_hashes:
                skipped_duplicates += 1
                continue
            exists = db.scalar(select(Transaction).where(Transaction.household_id == household_id, Transaction.tx_hash == tx_hash))
            if exists:
                skipped_duplicates += 1
                continue
            seen_hashes.add(tx_hash)

            db.add(
                Transaction(
                    household_id=household_id,
                    tx_date=tx_date,
                    tx_type=tx_type,
                    category=category,
                    subcategory=subcategory,
                    content=content,
                    amount=amount,
                    currency=currency,
                    payment_method=payment_method,
                    tx_hash=tx_hash,
                )
            )
            imported += 1
    else:
        # fallback simple format: date,type(asset/liability),name,amount
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1] or not row[2] or row[3] is None:
                continue
            as_of_date = excel_serial_to_date(row[0])
            kind = str(row[1]).lower().strip()
            name = str(row[2]).strip()
            amount = float(row[3])
            if kind == "asset":
                target = db.scalar(select(Asset).where(Asset.household_id == household_id, Asset.name == name))
                if not target:
                    target = Asset(household_id=household_id, name=name, category="imported")
                    db.add(target)
                    db.flush()
                db.add(Valuation(household_id=household_id, asset_id=target.id, as_of_date=as_of_date, amount=amount))
                imported += 1
            elif kind == "liability":
                target = db.scalar(select(Liability).where(Liability.household_id == household_id, Liability.name == name))
                if not target:
                    target = Liability(household_id=household_id, name=name)
                    db.add(target)
                    db.flush()
                db.add(Valuation(household_id=household_id, liability_id=target.id, as_of_date=as_of_date, amount=amount))
                imported += 1

    audit(
        db,
        household_id,
        user.id,
        "import",
        "xlsx",
        file.filename or "upload",
        detail=f"transactions={imported}, skipped_duplicates={skipped_duplicates}, assets={imported_assets}, liabilities={imported_liabilities}",
    )
    db.commit()
    return {
        "imported": imported,
        "skipped_duplicates": skipped_duplicates,
        "imported_assets": imported_assets,
        "imported_liabilities": imported_liabilities,
    }


@app.post("/imports/xlsx-local")
def import_xlsx_local(file: UploadFile = File(...), db: Session = Depends(get_db)):
    demo, household = ensure_local_household(db)
    return import_xlsx(household_id=household.id, file=file, user=demo, db=db)


@app.post("/snapshots/recompute")
def recompute_snapshots(household_id: int = 1, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "member")

    # 재계산 시 기존 스냅샷을 지우고 한 번에 재생성(중복 충돌 방지)
    db.query(NetWorthSnapshot).filter(NetWorthSnapshot.household_id == household_id).delete()
    db.flush()

    by_date: dict[date, dict[str, float]] = {}

    # A) 거래 기반 일자 누적 순자산(타임라인)
    tx_rows = db.execute(
        select(Transaction.tx_date, func.coalesce(func.sum(Transaction.amount), 0))
        .where(Transaction.household_id == household_id)
        .group_by(Transaction.tx_date)
        .order_by(Transaction.tx_date.asc())
    ).all()

    running = 0.0
    for d, day_sum in tx_rows:
        d = norm_date(d)
        running += float(day_sum)
        by_date[d] = {
            "assets_total": running,
            "liabilities_total": 0.0,
            "net_worth": running,
        }

    # B) valuation(시트1) 값은 해당 날짜 스냅샷에 덮어써 최신 자산/부채 기준 반영
    val_dates = db.scalars(select(Valuation.as_of_date).where(Valuation.household_id == household_id).distinct()).all()
    for d in val_dates:
        d = norm_date(d)
        assets_total = float(db.scalar(
            select(func.coalesce(func.sum(Valuation.amount), 0)).where(
                Valuation.household_id == household_id,
                Valuation.as_of_date == d,
                Valuation.asset_id.is_not(None),
            )
        ) or 0)
        liabilities_total = float(db.scalar(
            select(func.coalesce(func.sum(Valuation.amount), 0)).where(
                Valuation.household_id == household_id,
                Valuation.as_of_date == d,
                Valuation.liability_id.is_not(None),
            )
        ) or 0)
        net = assets_total - liabilities_total
        by_date[d] = {
            "assets_total": assets_total,
            "liabilities_total": liabilities_total,
            "net_worth": net,
        }

    for d in sorted(by_date.keys()):
        item = by_date[d]
        db.add(NetWorthSnapshot(
            household_id=household_id,
            snapshot_date=d,
            assets_total=item["assets_total"],
            liabilities_total=item["liabilities_total"],
            net_worth=item["net_worth"],
        ))

    count = len(by_date)
    audit(db, household_id, user.id, "recompute", "net_worth_snapshots", str(count))
    db.commit()
    return {"snapshots": count}


@app.get("/households/{household_id}/reports/monthly")
def monthly_report(
    household_id: int,
    year: int,
    month: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_household_role(db, household_id, user.id, "viewer")
    ym = f"{year:04d}-{month:02d}"

    income = db.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            Transaction.household_id == household_id,
            Transaction.tx_type == "수입",
            func.strftime("%Y-%m", Transaction.tx_date) == ym,
        )
    )
    expense = db.scalar(
        select(func.coalesce(func.sum(func.abs(Transaction.amount)), 0)).where(
            Transaction.household_id == household_id,
            Transaction.tx_type == "지출",
            func.strftime("%Y-%m", Transaction.tx_date) == ym,
        )
    )

    by_category_rows = db.execute(
        select(Transaction.category, func.coalesce(func.sum(Transaction.amount), 0))
        .where(
            Transaction.household_id == household_id,
            Transaction.tx_type == "지출",
            func.strftime("%Y-%m", Transaction.tx_date) == ym,
        )
        .group_by(Transaction.category)
        .order_by(func.sum(Transaction.amount).asc())
    ).all()

    by_category = [
        {"category": r[0] or "미분류", "amount": abs(float(r[1]))}
        for r in by_category_rows
    ]

    return {
        "year": year,
        "month": month,
        "income": float(income),
        "expense": float(expense),
        "cashflow": float(income) - float(expense),
        "expenseByCategory": by_category,
    }


@app.get("/households/{household_id}/category-share")
def category_share(
    household_id: int,
    year: int,
    month: int,
    tx_type: str = "지출",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_household_role(db, household_id, user.id, "viewer")
    ym = f"{year:04d}-{month:02d}"

    rows = db.execute(
        select(Transaction.category, func.coalesce(func.sum(Transaction.amount), 0))
        .where(
            Transaction.household_id == household_id,
            Transaction.tx_type == tx_type,
            func.strftime("%Y-%m", Transaction.tx_date) == ym,
        )
        .group_by(Transaction.category)
    ).all()

    parsed = []
    for c, amt in rows:
        v = abs(float(amt))
        if v <= 0:
            continue
        parsed.append({"category": c or "미분류", "amount": v})

    parsed.sort(key=lambda x: x["amount"], reverse=True)
    total = sum(x["amount"] for x in parsed)
    out = []
    for x in parsed:
        weight = (x["amount"] / total * 100) if total > 0 else 0
        out.append({**x, "weight": round(weight, 2)})

    return {"tx_type": tx_type, "total": total, "items": out}


@app.get("/households/{household_id}/cashflow/daily")
def daily_cashflow(
    household_id: int,
    year: int,
    month: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_household_role(db, household_id, user.id, "viewer")
    ym = f"{year:04d}-{month:02d}"

    rows = db.execute(
        select(Transaction.tx_date, Transaction.tx_type, func.coalesce(func.sum(Transaction.amount), 0))
        .where(
            Transaction.household_id == household_id,
            func.strftime("%Y-%m", Transaction.tx_date) == ym,
        )
        .group_by(Transaction.tx_date, Transaction.tx_type)
        .order_by(Transaction.tx_date.asc())
    ).all()

    by_day: dict[str, dict[str, float]] = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "transfer": 0.0, "net": 0.0})
    for d, tx_type, amt in rows:
        key = norm_date(d).isoformat()
        v = float(amt)
        if tx_type == "수입":
            by_day[key]["income"] += v
            by_day[key]["net"] += v
        elif tx_type == "지출":
            by_day[key]["expense"] += abs(v)
            by_day[key]["net"] -= abs(v)
        else:
            by_day[key]["transfer"] += v

    return [{"date": k, **by_day[k]} for k in sorted(by_day.keys())]


@app.get("/households/{household_id}/cashflow/daily-report")
def daily_report(
    household_id: int,
    day: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_household_role(db, household_id, user.id, "viewer")
    target = norm_date(day)

    rows = db.execute(
        select(Transaction.tx_type, Transaction.category, func.coalesce(func.sum(Transaction.amount), 0))
        .where(Transaction.household_id == household_id, Transaction.tx_date == target)
        .group_by(Transaction.tx_type, Transaction.category)
    ).all()

    income = 0.0
    expense = 0.0
    by_cat = []
    for tx_type, cat, amt in rows:
        v = float(amt)
        if tx_type == "수입":
            income += v
        elif tx_type == "지출":
            expense += abs(v)
        by_cat.append({"type": tx_type, "category": cat or "미분류", "amount": abs(v)})

    txs = db.execute(
        select(Transaction.tx_type, Transaction.category, Transaction.content, Transaction.amount, Transaction.payment_method)
        .where(Transaction.household_id == household_id, Transaction.tx_date == target)
        .order_by(func.abs(Transaction.amount).desc())
        .limit(30)
    ).all()

    tx_list = [
        {
            "type": t[0],
            "category": t[1] or "미분류",
            "content": t[2] or "",
            "amount": float(t[3]),
            "paymentMethod": t[4] or "",
        }
        for t in txs
    ]

    return {
        "date": target.isoformat(),
        "income": income,
        "expense": expense,
        "net": income - expense,
        "categories": by_cat,
        "transactions": tx_list,
    }


@app.get("/households/{household_id}/cashflow/monthly")
def monthly_cashflow(household_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "viewer")
    rows = db.execute(
        select(
            func.strftime("%Y-%m", Transaction.tx_date).label("ym"),
            Transaction.tx_type,
            func.coalesce(func.sum(Transaction.amount), 0),
        )
        .where(Transaction.household_id == household_id)
        .group_by(func.strftime("%Y-%m", Transaction.tx_date), Transaction.tx_type)
        .order_by(func.strftime("%Y-%m", Transaction.tx_date).asc())
    ).all()

    agg: dict[str, dict[str, float]] = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "transfer": 0.0})
    for ym, tx_type, amt in rows:
        if tx_type == "수입":
            agg[ym]["income"] += float(amt)
        elif tx_type == "지출":
            agg[ym]["expense"] += abs(float(amt))
        else:
            agg[ym]["transfer"] += float(amt)

    out = []
    for ym in sorted(agg.keys()):
        item = agg[ym]
        out.append({
            "month": ym,
            "income": item["income"],
            "expense": item["expense"],
            "transfer": item["transfer"],
            "cashflow": item["income"] - item["expense"],
        })
    return out


@app.get("/households/{household_id}/flow")
def flow_chart(household_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "viewer")

    rows = db.execute(
        select(Transaction.tx_type, Transaction.category, func.coalesce(func.sum(Transaction.amount), 0))
        .where(Transaction.household_id == household_id)
        .group_by(Transaction.tx_type, Transaction.category)
    ).all()

    incomes: list[tuple[str, float]] = []
    expenses: list[tuple[str, float]] = []
    transfer_total = 0.0

    for tx_type, category, amt in rows:
        v = abs(float(amt))
        if v <= 0:
            continue
        c = category or "미분류"
        if tx_type == "수입":
            incomes.append((c, v))
        elif tx_type == "지출":
            expenses.append((c, v))
        else:
            transfer_total += v

    def top_n(items: list[tuple[str, float]], n: int = 8) -> list[tuple[str, float]]:
        items = sorted(items, key=lambda x: x[1], reverse=True)
        top = items[:n]
        remain = sum(v for _, v in items[n:])
        if remain > 0:
            top.append(("기타", remain))
        return top

    incomes = top_n(incomes)
    expenses = top_n(expenses)

    total_income = sum(v for _, v in incomes)
    total_expense = sum(v for _, v in expenses)
    net = total_income - total_expense

    nodes = [
        {"name": "총수입"},
        {"name": "총지출"},
        {"name": "순현금흐름"},
        {"name": "흑자" if net >= 0 else "적자"},
    ]
    idx = {n["name"]: i for i, n in enumerate(nodes)}
    links = []

    for cat, v in incomes:
        nm = f"수입·{cat}"
        idx[nm] = len(nodes)
        nodes.append({"name": nm})
        links.append({"source": idx[nm], "target": idx["총수입"], "value": v})

    for cat, v in expenses:
        nm = f"지출·{cat}"
        idx[nm] = len(nodes)
        nodes.append({"name": nm})
        links.append({"source": idx["총지출"], "target": idx[nm], "value": v})

    if total_income > 0:
        links.append({"source": idx["총수입"], "target": idx["순현금흐름"], "value": total_income})
    if total_expense > 0:
        links.append({"source": idx["순현금흐름"], "target": idx["총지출"], "value": total_expense})
    if abs(net) > 0:
        links.append({"source": idx["순현금흐름"], "target": idx["흑자" if net >= 0 else "적자"], "value": abs(net)})
    if transfer_total > 0:
        idx["이체"] = len(nodes)
        nodes.append({"name": "이체"})
        links.append({"source": idx["이체"], "target": idx["순현금흐름"], "value": transfer_total})

    return {
        "nodes": nodes,
        "links": links,
        "summary": {"income": total_income, "expense": total_expense, "net": net, "transfer": transfer_total},
        "incomeBreakdown": [{"category": c, "amount": v} for c, v in incomes],
        "expenseBreakdown": [{"category": c, "amount": v} for c, v in expenses],
    }


@app.get("/households/{household_id}/balance-sheet")
def balance_sheet(household_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "viewer")

    latest_date = db.scalar(select(func.max(Valuation.as_of_date)).where(Valuation.household_id == household_id))
    if not latest_date:
        return {"asOf": None, "assetsTotal": 0, "liabilitiesTotal": 0, "assets": [], "liabilities": []}

    asset_rows = db.execute(
        select(Asset.name, func.coalesce(func.sum(Valuation.amount), 0))
        .join(Asset, Asset.id == Valuation.asset_id)
        .where(Valuation.household_id == household_id, Valuation.as_of_date == latest_date)
        .group_by(Asset.name)
        .order_by(func.sum(Valuation.amount).desc())
    ).all()

    liability_rows = db.execute(
        select(Liability.name, func.coalesce(func.sum(Valuation.amount), 0))
        .join(Liability, Liability.id == Valuation.liability_id)
        .where(Valuation.household_id == household_id, Valuation.as_of_date == latest_date)
        .group_by(Liability.name)
        .order_by(func.sum(Valuation.amount).desc())
    ).all()

    assets_total = sum(float(v) for _, v in asset_rows)
    liabilities_total = sum(float(v) for _, v in liability_rows)

    assets = [
        {"name": n, "value": float(v), "weight": round((float(v) / assets_total * 100), 2) if assets_total > 0 else 0}
        for n, v in asset_rows
    ]
    liabilities = [
        {"name": n, "value": float(v), "weight": round((float(v) / liabilities_total * 100), 2) if liabilities_total > 0 else 0}
        for n, v in liability_rows
    ]

    return {
        "asOf": latest_date.isoformat(),
        "assetsTotal": assets_total,
        "liabilitiesTotal": liabilities_total,
        "assets": assets,
        "liabilities": liabilities,
    }


@app.get("/households/{household_id}/net-worth")
def net_worth(household_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    require_household_role(db, household_id, user.id, "viewer")
    rows = db.scalars(
        select(NetWorthSnapshot).where(NetWorthSnapshot.household_id == household_id).order_by(NetWorthSnapshot.snapshot_date.asc())
    ).all()
    return [
        {
            "date": r.snapshot_date.isoformat(),
            "assets": float(r.assets_total),
            "liabilities": float(r.liabilities_total),
            "netWorth": float(r.net_worth),
        }
        for r in rows
    ]
